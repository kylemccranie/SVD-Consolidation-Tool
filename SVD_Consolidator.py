import os
from os import path
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill


print("Enter the file path for the folder that cotains the SVDs you wish to consolidate")
basepath = input()
os.chdir(basepath)

WD = os.listdir(basepath)


df = pd.DataFrame()
for i in WD:
    os.chdir(basepath + "/" + i)
    if os.path.exists("SoftwareChanges.xlsx") == True:
        df = df.append(pd.read_excel('SoftwareChanges.xlsx', skiprows = range(2), header=None), ignore_index= True)

os.chdir(basepath)

df.to_excel('Consolidated_MPE_SVD.xlsx')


def Format():
    wb = load_workbook(filename='Consolidated_MPE_SVD.xlsx')
    ws = wb.active
    ws.delete_cols(1)
    ws.delete_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.merge_cells('A1:M1')
    ws.cell(1,1).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,1).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,2).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,3).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,4).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,5).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,6).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,7).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,8).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,9).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,10).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,11).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,12).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)
    ws.cell(2,13).alignment = Alignment(horizontal='center', vertical= 'center',wrap_text=True)

    font_black = Font(name='Calibri',
                        size=12,
                        bold=True,
                        color='000000')
    title = Font(name='Calibri',
                        size=24,
                        bold=True,
                        color='000000')
    ws['A1'].font = title                    
    ws['A2'].font = font_black
    ws['B2'].font = font_black 
    ws['C2'].font = font_black 
    ws['D2'].font = font_black
    ws['E2'].font = font_black 
    ws['F2'].font = font_black
    ws['G2'].font = font_black
    ws['H2'].font = font_black
    ws['I2'].font = font_black 
    ws['J2'].font = font_black
    ws['K2'].font = font_black 
    ws['L2'].font = font_black
    ws['M2'].font = font_black
    ws.row_dimensions[1].height = 37
    ws.row_dimensions[2].height = 60
    ws.column_dimensions['A'].width = 11.86
    ws.column_dimensions['B'].width = 11.29
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12.71
    ws.column_dimensions['F'].width = 19
    ws.column_dimensions['G'].width = 19
    ws.column_dimensions['H'].width = 19
    ws.column_dimensions['I'].width = 19
    ws.column_dimensions['J'].width = 8.57
    ws.column_dimensions['K'].width = 19
    ws.column_dimensions['L'].width = 19
    ws.column_dimensions['M'].width = 19
    ws['A1'] = "Software Changes"
    ws['A2'] = "Type"
    ws['B2'] = "DR Number"
    ws['C2'] = "Internal ID"
    ws['D2'] = "Component Variant"
    ws['E2'] = "CVV Build #"
    ws['F2'] = "Title"
    ws['G2'] = "Description"
    ws['H2'] = "Problem Path"
    ws['I2'] = "Acceptance Criteria"
    ws['J2'] = "UI Affected"
    ws['K2'] = "If UI is Affected Then How"
    ws['L2'] = "Change Summary"
    ws['M2'] = "Fix Comments"
    ws.freeze_panes = "Z3"
    wb.save('Consolidated_MPE_SVD_final.xlsx')

Format()
